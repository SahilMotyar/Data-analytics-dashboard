from __future__ import annotations

import asyncio
import hashlib
import json
import logging
import threading
import traceback
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from fastapi import APIRouter, File, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel

from app.core.config import settings
from app.core.storage import read_json, write_json
from app.services.analysis_service import (
    analysis_path,
    get_analysis,
    metadata_path,
    set_column_type_override,
    upload_dir,
)
from app.services.export_service import create_pdf_report, create_share_link
from app.services.ai_service import generate_chat_answer
from app.workers.tasks import run_analysis_pipeline

logger = logging.getLogger(__name__)


router = APIRouter(prefix="/api/v1")


def _rate_limit_path() -> Path:
    return settings.storage_dir / "rate_limit.json"


def _apply_upload_rate_limit(client_key: str) -> None:
    store = read_json(_rate_limit_path(), {})
    day_key = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    client_day_key = f"{client_key}:{day_key}"
    count = int(store.get(client_day_key, 0))

    if count >= settings.upload_daily_limit:
        raise HTTPException(status_code=429, detail="Upload daily limit reached (50/day)")

    store[client_day_key] = count + 1
    write_json(_rate_limit_path(), store)


class StartAnalysisRequest(BaseModel):
    active_sheet: str | None = None
    mode: str = "auto"           # "auto" | "quick" | "focused" | "full"
    focus_columns: list[str] = []  # used with mode="focused"


class OverrideTypeRequest(BaseModel):
    new_type: str


class ChatRequest(BaseModel):
    message: str
    history: list[dict[str, str]] = []


@router.get("/analysis/{upload_id}/grid-preview")
def grid_preview(
    upload_id: str,
    limit: int = 200,
    outliers_only: bool = False,
    missing_only: bool = False,
) -> dict[str, Any]:
    analysis = get_analysis(upload_id)
    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found")

    cleaned_path = settings.outputs_dir / upload_id / "cleaned.csv"
    if not cleaned_path.exists():
        raise HTTPException(status_code=404, detail="Cleaned data not found")

    df = pd.read_csv(cleaned_path)
    pre = analysis.get("pre_analysis", {})
    smart = pre.get("smart_type_correction", {})
    reclassified = {item.get("column") for item in smart.get("reclassifications", [])}
    outlier_map = pre.get("outlier_characterisation", {}).get("outliers_by_column", {})
    anomaly_rows = {
        int(item.get("row_index"))
        for item in pre.get("outlier_characterisation", {}).get("multi_column_anomalies", [])
    }

    outlier_row_col = {
        (int(detail.get("row_index")), col)
        for col, payload in outlier_map.items()
        for detail in payload.get("details", [])
    }

    if outliers_only:
        outlier_rows = sorted({row for row, _ in outlier_row_col})
        df = df.loc[df.index.intersection(outlier_rows)]
    if missing_only:
        df = df[df.isna().any(axis=1)]

    df = df.head(max(1, min(limit, 2000)))
    rows = []
    for idx, row in df.iterrows():
        values = row.to_dict()
        cell_flags = {}
        for col, value in values.items():
            flags = []
            if pd.isna(value):
                flags.append("missing")
            if (int(idx), col) in outlier_row_col:
                flags.append("outlier")
            if col in reclassified:
                flags.append("reclassified")
            cell_flags[col] = flags
        rows.append(
            {
                "row_index": int(idx),
                "values": values,
                "cell_flags": cell_flags,
                "row_flags": ["multi_column_anomaly"] if int(idx) in anomaly_rows else [],
            }
        )

    columns_payload = []
    for col in df.columns:
        stats = analysis.get("column_stats", {}).get(col, {})
        final_type = stats.get("inferred_type")
        rec_item = next((item for item in smart.get("reclassifications", []) if item.get("column") == col), None)
        missing_pct = stats.get("missing_pct", 0)
        mini_dist = {
            "kind": "numeric" if final_type == "numeric" else "categorical" if final_type in {"categorical", "boolean"} else "datetime" if final_type == "datetime" else "other",
            "shape": (stats.get("distribution_shape") or {}).get("shape"),
            "top_10": stats.get("top_10", [])[:5],
        }
        columns_payload.append(
            {
                "name": col,
                "final_type": final_type,
                "reclassified": rec_item is not None,
                "reclassification": rec_item,
                "missing_pct": missing_pct,
                "mini_distribution": mini_dist,
            }
        )

    return {
        "upload_id": upload_id,
        "rows": rows,
        "columns": columns_payload,
        "applied_filters": {"outliers_only": outliers_only, "missing_only": missing_only, "limit": limit},
    }


@router.post("/uploads")
async def upload_file(request: Request, file: UploadFile = File(...)) -> dict[str, Any]:
    client_key = request.client.host if request.client else "anonymous"
    _apply_upload_rate_limit(client_key)

    suffix = Path(file.filename).suffix.lower()
    if suffix not in {".csv", ".xlsx", ".xls"}:
        raise HTTPException(status_code=400, detail="Only CSV/XLS/XLSX files are supported")

    contents = await file.read()
    size_mb = len(contents) / (1024 * 1024)
    if size_mb > settings.max_upload_mb:
        raise HTTPException(status_code=413, detail="Max file size is 50MB")

    upload_id = str(uuid.uuid4())
    target_dir = upload_dir(upload_id) / "original"
    target_dir.mkdir(parents=True, exist_ok=True)

    file_path = target_dir / file.filename
    with file_path.open("wb") as output:
        output.write(contents)

    sheet_names = []
    active_sheet = None
    if suffix in {".xlsx", ".xls"}:
        excel = pd.ExcelFile(file_path)
        sheet_names = excel.sheet_names
        active_sheet = sheet_names[0] if sheet_names else None

    warning = "Large file (>100K rows) may take longer to process"
    row_warning = None
    try:
        if suffix == ".csv":
            sample_rows = sum(1 for _ in file_path.open("r", encoding="utf-8", errors="ignore")) - 1
        else:
            sample_rows = int(pd.read_excel(file_path, sheet_name=active_sheet).shape[0])
        if sample_rows > 100_000:
            row_warning = warning
    except Exception:
        sample_rows = None

    metadata = {
        "upload_id": upload_id,
        "file_name": file.filename,
        "file_hash": hashlib.sha256(contents).hexdigest(),
        "file_size_mb": round(size_mb, 2),
        "status": "uploaded",
        "analysis_status": "not_started",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "sheet_names": sheet_names,
        "active_sheet": active_sheet,
        "row_count_estimate": sample_rows,
        "warning": row_warning,
    }
    write_json(metadata_path(upload_id), metadata)

    return {"upload_id": upload_id, "metadata": metadata}


@router.get("/uploads/{upload_id}")
def get_upload(upload_id: str) -> dict[str, Any]:
    meta = read_json(metadata_path(upload_id), None)
    if not meta:
        raise HTTPException(status_code=404, detail="Upload not found")

    # Detect column names quickly so frontend can show column picker
    if "detected_columns" not in meta:
        try:
            file_name = meta.get("file_name", "")
            file_path = upload_dir(upload_id) / "original" / file_name
            if file_name.lower().endswith(".csv"):
                cols = pd.read_csv(file_path, nrows=0).columns.tolist()
            elif file_name.lower().endswith((".xlsx", ".xls")):
                sheet = meta.get("active_sheet")
                cols = pd.read_excel(file_path, sheet_name=sheet, nrows=0).columns.tolist()
            else:
                cols = []
            meta["detected_columns"] = cols
            write_json(metadata_path(upload_id), meta)
        except Exception:
            meta["detected_columns"] = []

    return meta


def _run_pipeline_in_thread(upload_id: str) -> None:
    """Run the analysis pipeline in a background thread so /start returns immediately."""
    try:
        from app.core.storage import read_json as _rj, write_json as _wj
        from app.services.analysis_service import metadata_path as _mp, analysis_path as _ap, compute_analysis
        from app.services.ai_service import generate_ai_summaries

        meta = _rj(_mp(upload_id), {})
        meta["analysis_status"] = "running"
        _wj(_mp(upload_id), meta)

        mode = meta.get("analysis_mode", "auto")
        focus_columns = meta.get("focus_columns", [])

        analysis = compute_analysis(
            upload_id,
            sheet_name=meta.get("active_sheet"),
            mode=mode,
            focus_columns=focus_columns,
        )
        analysis = generate_ai_summaries(analysis, meta)
        _wj(_ap(upload_id), analysis)

        meta = _rj(_mp(upload_id), {})
        meta["analysis_status"] = "completed"
        _wj(_mp(upload_id), meta)
        logger.info("Analysis completed for %s (mode=%s)", upload_id, mode)
    except Exception as exc:
        logger.error("Analysis failed for %s: %s", upload_id, traceback.format_exc())
        meta = _rj(_mp(upload_id), {})
        meta["analysis_status"] = "failed"
        meta["error"] = str(exc)
        _wj(_mp(upload_id), meta)


@router.post("/analysis/{upload_id}/start")
def start_analysis(upload_id: str, payload: StartAnalysisRequest) -> dict[str, Any]:
    meta = read_json(metadata_path(upload_id), None)
    if not meta:
        raise HTTPException(status_code=404, detail="Upload not found")

    if payload.active_sheet:
        meta["active_sheet"] = payload.active_sheet
    meta["analysis_mode"] = payload.mode
    meta["focus_columns"] = payload.focus_columns
    meta["analysis_status"] = "queued"
    write_json(metadata_path(upload_id), meta)

    if settings.celery_task_always_eager:
        # Run in a thread so the HTTP response returns instantly
        t = threading.Thread(target=_run_pipeline_in_thread, args=(upload_id,), daemon=True)
        t.start()
        return {"upload_id": upload_id, "job_id": "thread", "status": "queued"}

    job = run_analysis_pipeline.delay(upload_id)
    return {"upload_id": upload_id, "job_id": job.id, "status": "queued"}


@router.get("/analysis/{upload_id}/status")
def analysis_status(upload_id: str) -> dict[str, Any]:
    meta = read_json(metadata_path(upload_id), None)
    if not meta:
        raise HTTPException(status_code=404, detail="Upload not found")
    return {"upload_id": upload_id, "status": meta.get("analysis_status", "not_started"), "error": meta.get("error")}


@router.get("/analysis/{upload_id}/summary")
def analysis_summary(upload_id: str) -> dict[str, Any]:
    analysis = get_analysis(upload_id)
    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found")
    return analysis.get("summary", {})


@router.get("/analysis/{upload_id}/key-findings")
def key_findings(upload_id: str) -> dict[str, Any]:
    analysis = get_analysis(upload_id)
    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found")
    return analysis.get("key_findings", {})


@router.get("/analysis/{upload_id}/pre-analysis")
def pre_analysis(upload_id: str) -> dict[str, Any]:
    analysis = get_analysis(upload_id)
    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found")
    return analysis.get("pre_analysis", {})


@router.get("/analysis/{upload_id}/columns")
def analysis_columns(upload_id: str) -> list[dict[str, Any]]:
    analysis = get_analysis(upload_id)
    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found")
    return analysis.get("columns", [])


@router.get("/analysis/{upload_id}/columns/{col}/stats")
def column_stats(upload_id: str, col: str) -> dict[str, Any]:
    analysis = get_analysis(upload_id)
    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found")

    stats = analysis.get("column_stats", {}).get(col)
    if not stats:
        raise HTTPException(status_code=404, detail="Column not found")
    return stats


@router.get("/analysis/{upload_id}/compare")
def compare_columns(upload_id: str, col_a: str, col_b: str) -> dict[str, Any]:
    analysis = get_analysis(upload_id)
    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found")

    cleaned_path = settings.outputs_dir / upload_id / "cleaned.csv"
    if not cleaned_path.exists():
        raise HTTPException(status_code=404, detail="Cleaned CSV not available")

    df = pd.read_csv(cleaned_path)
    if col_a not in df.columns or col_b not in df.columns:
        raise HTTPException(status_code=404, detail="Column not found")

    type_a = analysis.get("column_stats", {}).get(col_a, {}).get("inferred_type")
    type_b = analysis.get("column_stats", {}).get(col_b, {}).get("inferred_type")

    numeric_types = {"numeric"}
    cat_types = {"categorical", "boolean"}

    # numeric + numeric -> scatter relationship stats
    if type_a in numeric_types and type_b in numeric_types:
        sub = df[[col_a, col_b]].copy()
        sub[col_a] = pd.to_numeric(sub[col_a], errors="coerce")
        sub[col_b] = pd.to_numeric(sub[col_b], errors="coerce")
        sub = sub.dropna()
        if len(sub) < 5:
            return {"mode": "numeric_numeric", "message": "Not enough paired data points."}
        pearson = float(sub[col_a].corr(sub[col_b], method="pearson"))
        spearman = float(sub[col_a].corr(sub[col_b], method="spearman"))
        points = [
            {"x": float(row[col_a]), "y": float(row[col_b])}
            for _, row in sub.head(4000).iterrows()
        ]
        return {
            "mode": "numeric_numeric",
            "col_a": col_a,
            "col_b": col_b,
            "pearson_r": round(pearson, 4),
            "spearman_rho": round(spearman, 4),
            "non_linear_signal": abs(pearson - spearman) > 0.2,
            "points": points,
            "interpretation": f"{col_a} and {col_b} show correlation r={round(pearson, 3)}.",
        }

    # numeric + categorical -> grouped boxplot payload + effect size
    if (type_a in numeric_types and type_b in cat_types) or (type_b in numeric_types and type_a in cat_types):
        num_col = col_a if type_a in numeric_types else col_b
        cat_col = col_b if num_col == col_a else col_a
        sub = df[[num_col, cat_col]].copy()
        sub[num_col] = pd.to_numeric(sub[num_col], errors="coerce")
        sub = sub.dropna()
        grouped = sub.groupby(cat_col)[num_col].agg(["mean", "std", "count", "min", "max"]).reset_index()
        overall_std = float(sub[num_col].std()) if len(sub) else 0.0
        effect_size = 0.0 if overall_std == 0 else float((grouped["mean"].max() - grouped["mean"].min()) / overall_std)
        return {
            "mode": "numeric_categorical",
            "numeric_column": num_col,
            "categorical_column": cat_col,
            "effect_size": round(effect_size, 4),
            "groups": [
                {
                    "group": str(row[cat_col]),
                    "mean": round(float(row["mean"]), 4),
                    "std": round(float(row["std"]), 4) if not pd.isna(row["std"]) else None,
                    "count": int(row["count"]),
                    "min": round(float(row["min"]), 4),
                    "max": round(float(row["max"]), 4),
                }
                for _, row in grouped.iterrows()
            ],
            "interpretation": f"{cat_col} groups separate {num_col} with effect size {round(effect_size, 3)}.",
        }

    # categorical + categorical -> crosstab heatmap payload
    if type_a in cat_types and type_b in cat_types:
        sub = df[[col_a, col_b]].copy().dropna()
        tab = pd.crosstab(sub[col_a].astype(str), sub[col_b].astype(str))
        records = []
        for idx_name, row in tab.iterrows():
            for col_name, value in row.items():
                records.append({"x": str(idx_name), "y": str(col_name), "count": int(value)})
        most_common = max(records, key=lambda item: item["count"]) if records else None
        rarest = min((item for item in records if item["count"] > 0), key=lambda item: item["count"], default=None)
        return {
            "mode": "categorical_categorical",
            "col_a": col_a,
            "col_b": col_b,
            "cells": records,
            "most_common": most_common,
            "rarest": rarest,
            "interpretation": f"Most common combination is {most_common} and rarest non-zero is {rarest}." if most_common else "No overlapping values.",
        }

    return {"mode": "unsupported", "message": "Unsupported column type combination."}


@router.patch("/analysis/{upload_id}/columns/{col}/type")
def override_column_type(upload_id: str, col: str, payload: OverrideTypeRequest) -> dict[str, Any]:
    try:
        stats = set_column_type_override(upload_id, col, payload.new_type)
        return {"column": col, "new_type": payload.new_type, "stats": stats}
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@router.post("/analysis/{upload_id}/export/pdf")
def export_pdf(upload_id: str) -> dict[str, Any]:
    analysis = get_analysis(upload_id)
    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found")
    pdf_url = create_pdf_report(upload_id, analysis)
    return {"pdf_url": pdf_url}


@router.get("/analysis/{upload_id}/export/cleaned-csv")
def export_cleaned_csv(upload_id: str):
    cleaned_path = settings.outputs_dir / upload_id / "cleaned.csv"
    if not cleaned_path.exists():
        raise HTTPException(status_code=404, detail="Cleaned CSV not available")
    return FileResponse(cleaned_path, media_type="text/csv", filename="cleaned.csv")


@router.get("/analysis/{upload_id}/export/excel")
def export_excel(upload_id: str, sample: int | None = None):
    """Export to Excel. Pass ?sample=10000 to only export the first N rows (fast for huge files)."""
    cleaned_path = settings.outputs_dir / upload_id / "cleaned.csv"
    if not cleaned_path.exists():
        raise HTTPException(status_code=404, detail="Cleaned CSV not available")

    suffix = f"_sample{sample}" if sample else ""
    excel_path = settings.outputs_dir / upload_id / f"cleaned{suffix}.xlsx"

    # For large files, use chunked reading + openpyxl write_only mode
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Data")

    rows_written = 0
    for chunk_idx, chunk in enumerate(pd.read_csv(cleaned_path, chunksize=10_000)):
        if chunk_idx == 0:
            ws.append(list(chunk.columns))
        for _, row in chunk.iterrows():
            ws.append([v if not pd.isna(v) else None for v in row.tolist()])
            rows_written += 1
            if sample and rows_written >= sample:
                break
        if sample and rows_written >= sample:
            break

    # Add a summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary.append(["DataLens Export Summary"])
    ws_summary.append([])
    ws_summary.append(["Rows exported", rows_written])
    analysis = get_analysis(upload_id)
    if analysis:
        summary = analysis.get("summary", {})
        ws_summary.append(["Total rows in dataset", summary.get("rows", "?")])
        ws_summary.append(["Total columns", summary.get("columns", "?")])
        ws_summary.append(["Quality score", summary.get("quality_score", "?")])
        ws_summary.append([])
        ws_summary.append(["Column", "Type", "Missing %", "Health"])
        for col_info in analysis.get("columns", []):
            col_name = col_info["name"]
            col_type = col_info["inferred_type"]
            col_missing = col_info.get("quality_flags", {}).get("missing_pct", 0)
            col_health = col_info.get("health", {}).get("overall", {}).get("label", "OK")
            ws_summary.append([col_name, col_type, col_missing, col_health])

    wb.save(excel_path)

    filename = f"cleaned{suffix}.xlsx"
    return FileResponse(excel_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=filename)


@router.get("/analysis/{upload_id}/export/stats-json")
def export_stats_json(upload_id: str):
    analysis = get_analysis(upload_id)
    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found")
    stats_path = settings.outputs_dir / upload_id / "stats.json"
    write_json(stats_path, analysis)
    return FileResponse(stats_path, media_type="application/json", filename="stats.json")


@router.post("/analysis/{upload_id}/share")
def share_analysis(upload_id: str) -> dict[str, Any]:
    analysis = get_analysis(upload_id)
    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found")
    return create_share_link(upload_id)


@router.post("/analysis/{upload_id}/chat/stream")
async def chat_stream(upload_id: str, payload: ChatRequest):
    analysis = get_analysis(upload_id)
    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found")
    meta = read_json(metadata_path(upload_id), {})

    answer = generate_chat_answer(analysis, meta, payload.message, payload.history)

    async def event_generator():
        for token in answer.split(" "):
            yield f"data: {json.dumps({'token': token + ' '})}\n\n"
            await asyncio.sleep(0.01)
        yield f"data: {json.dumps({'done': True})}\n\n"

    return StreamingResponse(event_generator(), media_type="text/event-stream")


@router.get("/shared/{token}")
def get_shared(token: str) -> dict[str, Any]:
    from app.services.export_service import serializer

    try:
        payload = serializer.loads(token, max_age=7 * 24 * 3600)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Invalid/expired share token") from exc

    upload_id = payload["upload_id"]
    analysis = get_analysis(upload_id)
    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found")
    return {"upload_id": upload_id, "analysis": analysis}
